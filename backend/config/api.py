"""
NinjaAPI instance and project-level endpoints.
"""

import io
import uuid

from django.db import transaction
from django.http import HttpResponse, JsonResponse
from ninja import File, NinjaAPI, Schema, UploadedFile
from ninja.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scheduler import excel_parser, templates, validator
from scheduler.accounts import hash_password, mint_token, verify_password
from scheduler.auth import tenant_auth
from scheduler.models import (
    Module,
    Resource,
    Schedule,
    Session,
    SolveJob,
    StudentGroup,
    Teacher,
    TeacherModule,
    Tenant,
    User,
)
from scheduler.scoping import scoped_queryset
from scheduler.seed import seed_demo_data
from scheduler.solver import objectives


api = NinjaAPI(title="evo-scheduler-service API", version="1.0.0")


@api.get("/health")
def health(request):
    return {"status": "ok"}


class AuthUserOut(Schema):
    id: str
    email: str
    name: str


class AuthResponse(Schema):
    token: str
    user: AuthUserOut
    tenant_id: int


class RegisterIn(Schema):
    name: str
    email: str
    password: str


class LoginIn(Schema):
    email: str
    password: str


def _normalize_email(email: str) -> str:
    return (email or "").strip().lower()


def _user_payload(user: User) -> AuthUserOut:
    return {"id": str(user.id), "email": user.email, "name": user.name}


@api.post("/auth/register", response={201: AuthResponse})
def auth_register(request, payload: RegisterIn):
    name = (payload.name or "").strip()
    email = _normalize_email(payload.email)
    password = payload.password or ""
    if not name or not email or not password:
        raise HttpError(400, "name, email and password are required")
    if User.objects.filter(email=email).exists():
        raise HttpError(400, "Email already registered")
    with transaction.atomic():
        code_base = email.split("@")[0][:32] or "tenant"
        tenant_code = f"{code_base}-{uuid.uuid4().hex[:8]}"
        tenant = Tenant.objects.create(
            name=f"{name}'s School",
            code=tenant_code,
        )
        user = User.objects.create(
            email=email,
            name=name,
            password_hash=hash_password(password),
            tenant=tenant,
        )
        seed_demo_data(tenant)
    token = mint_token(user)
    return {"token": token, "user": _user_payload(user), "tenant_id": tenant.id}


@api.post("/auth/login", response=AuthResponse)
def auth_login(request, payload: LoginIn):
    email = _normalize_email(payload.email)
    password = payload.password or ""
    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        raise HttpError(401, "Invalid credentials")
    if not verify_password(password, user.password_hash):
        raise HttpError(401, "Invalid credentials")
    token = mint_token(user)
    return {
        "token": token,
        "user": _user_payload(user),
        "tenant_id": user.tenant_id,
    }


@api.get("/auth/me", auth=tenant_auth)
def auth_me(request):
    auth_payload = request.auth
    claims = auth_payload.get("claims", {})
    return {
        "tenant_id": auth_payload["tenant_id"],
        "deployment_id": auth_payload.get("deployment_id"),
        "email": claims.get("email"),
        "name": claims.get("name"),
        "claims": claims,
    }


class MeResponse(Schema):
    tenant_id: int
    deployment_id: str | None = None
    tenant_code: str
    tenant_name: str
    claims: dict


@api.get("/me", auth=tenant_auth, response=MeResponse)
def me(request):
    auth_payload = request.auth
    tenant = auth_payload["tenant"]
    return {
        "tenant_id": tenant.id,
        "deployment_id": auth_payload.get("deployment_id"),
        "tenant_code": tenant.code,
        "tenant_name": tenant.name,
        "claims": auth_payload.get("claims", {}),
    }


class TeacherOut(Schema):
    code: str
    name: str


@api.get("/teachers", auth=tenant_auth, response=list[TeacherOut])
def list_teachers(request):
    qs = scoped_queryset(Teacher)
    return [{"code": t.code, "name": t.name} for t in qs]


class WeightsIn(Schema):
    weights: dict


class WeightsOut(Schema):
    schedule_id: int
    weights: dict


def _get_tenant_schedule(request, schedule_id):
    try:
        schedule = Schedule.objects.get(id=schedule_id)
    except (Schedule.DoesNotExist, ValueError, TypeError):
        raise HttpError(404, "schedule not found")
    if schedule.tenant_id != request.auth["tenant_id"]:
        raise HttpError(404, "schedule not found")
    return schedule


class ScheduleOut(Schema):
    id: int
    name: str
    status: str
    tier: str | None = None
    objective_value: float | None = None


class ScheduleCreateIn(Schema):
    name: str
    tier: str | None = None


@api.get("/schedule", auth=tenant_auth, response=list[ScheduleOut])
def list_schedules(request):
    qs = scoped_queryset(Schedule)
    return [
        {
            "id": s.id,
            "name": s.name,
            "status": s.status,
            "tier": s.tier,
            "objective_value": s.objective_value,
        }
        for s in qs
    ]


@api.post("/schedule", auth=tenant_auth, response=ScheduleOut)
def create_schedule(request, payload: ScheduleCreateIn):
    tenant = request.auth["tenant"]
    if payload.tier and payload.tier not in dict(Schedule.Tier.choices):
        raise HttpError(400, "invalid tier")
    schedule = Schedule.objects.create(
        tenant=tenant,
        name=payload.name,
        tier=payload.tier,
        status=Schedule.Status.DRAFT,
    )
    return {
        "id": schedule.id,
        "name": schedule.name,
        "status": schedule.status,
        "tier": schedule.tier,
        "objective_value": schedule.objective_value,
    }


@api.get("/schedule/{schedule_id}/sessions", auth=tenant_auth)
def list_schedule_sessions(request, schedule_id):
    schedule = _get_tenant_schedule(request, schedule_id)
    tenant = request.auth["tenant"]
    qs = (
        tenant.sessions.select_related(
            "module", "student_group", "assigned_resource"
        )
        .prefetch_related("assigned_teachers")
    )
    rows = []
    for s in qs:
        ts = s.assigned_timeslot or {}
        rows.append(
            {
                "id": s.id,
                "module_code": s.module.code,
                "module_name": s.module.name,
                "student_group_code": s.student_group.code,
                "student_group_name": s.student_group.name,
                "session_type": s.session_type,
                "tier": s.tier,
                "duration_slots": s.duration_slots,
                "is_locked": s.is_locked,
                "day": ts.get("day"),
                "period": ts.get("period"),
                "day_name": ts.get("day_name"),
                "resource_code": (
                    s.assigned_resource.code if s.assigned_resource else None
                ),
                "teachers": [
                    {"code": t.code, "name": t.name}
                    for t in s.assigned_teachers.all()
                ],
            }
        )
    return {"schedule_id": schedule.id, "sessions": rows}


DAY_NAME_VI = {
    0: "Thứ 2",
    1: "Thứ 3",
    2: "Thứ 4",
    3: "Thứ 5",
    4: "Thứ 6",
    5: "Thứ 7",
    6: "Chủ nhật",
}


@api.get("/schedule/{schedule_id}/conflicts", auth=tenant_auth)
def list_schedule_conflicts(request, schedule_id):
    schedule = _get_tenant_schedule(request, schedule_id)
    tenant = request.auth["tenant"]
    qs = (
        tenant.sessions.select_related("module", "student_group", "assigned_resource")
        .prefetch_related("assigned_teachers")
    )
    buckets = {}
    for s in qs:
        ts = s.assigned_timeslot
        if not ts:
            continue
        day = ts.get("day")
        period = ts.get("period")
        if day is None or period is None:
            continue
        buckets.setdefault((day, period), []).append(s)

    conflicts = []
    for (day, period), group in buckets.items():
        n = len(group)
        for i in range(n):
            for j in range(i + 1, n):
                a = group[i]
                b = group[j]
                a_teachers = {t.code for t in a.assigned_teachers.all()}
                b_teachers = {t.code for t in b.assigned_teachers.all()}
                shared_teachers = a_teachers & b_teachers
                a_room = a.assigned_resource_id
                b_room = b.assigned_resource_id
                room_clash = (
                    a_room is not None and a_room == b_room
                )
                group_clash = a.student_group_id == b.student_group_id
                for kind, hit in (
                    ("teacher", bool(shared_teachers)),
                    ("room", room_clash),
                    ("student_group", group_clash),
                ):
                    if not hit:
                        continue
                    detail = None
                    if kind == "teacher":
                        detail = ", ".join(sorted(shared_teachers))
                    elif kind == "room":
                        detail = (
                            a.assigned_resource.code if a.assigned_resource else ""
                        )
                    elif kind == "student_group":
                        detail = a.student_group.code
                    conflicts.append(
                        {
                            "type": kind,
                            "session_ids": [a.id, b.id],
                            "day": day,
                            "period": period,
                            "day_name": DAY_NAME_VI.get(day, ts.get("day_name", "")),
                            "detail": detail,
                        }
                    )
    return {"schedule_id": schedule.id, "conflicts": conflicts}


EXPORT_HEADER_FILL = PatternFill(
    start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid"
)
EXPORT_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
EXPORT_WRAP = Alignment(wrap_text=True, vertical="top")


def _build_export_xlsx(tenant, schedule):
    qs = (
        tenant.sessions.select_related(
            "module", "student_group", "assigned_resource"
        )
        .prefetch_related("assigned_teachers")
    )
    assigned = []
    days_seen = set()
    periods_seen = set()
    for s in qs:
        ts = s.assigned_timeslot
        if not ts or ts.get("day") is None or ts.get("period") is None:
            continue
        days_seen.add(ts["day"])
        periods_seen.add(ts["period"])
        teacher_names = ", ".join(t.name for t in s.assigned_teachers.all())
        assigned.append(
            {
                "day": ts["day"],
                "period": ts["period"],
                "module_code": s.module.code,
                "module_name": s.module.name,
                "group": s.student_group.code,
                "teachers": teacher_names,
                "room": s.assigned_resource.code if s.assigned_resource else "",
                "session_type": s.session_type,
            }
        )

    day_list = sorted(days_seen) if days_seen else list(range(0, 6))
    period_list = sorted(periods_seen) if periods_seen else list(range(0, 5))

    cell_map = {}
    for a in assigned:
        cell_map.setdefault((a["day"], a["period"]), []).append(a)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lich"
    ws.cell(row=1, column=1, value="Tiết / Ngày")
    for col_idx, d in enumerate(day_list, start=2):
        ws.cell(row=1, column=col_idx, value=DAY_NAME_VI.get(d, "Ngày %s" % d))
    for r_idx, p in enumerate(period_list, start=2):
        ws.cell(row=r_idx, column=1, value="Tiết %s" % (p + 1))
        for c_idx, d in enumerate(day_list, start=2):
            items = cell_map.get((d, p), [])
            if not items:
                continue
            lines = []
            for it in items:
                lines.append(
                    "%s - %s\n%s | %s | %s"
                    % (
                        it["module_code"],
                        it["group"],
                        it["teachers"] or "(chưa phân GV)",
                        it["room"] or "(chưa phân phòng)",
                        it["session_type"],
                    )
                )
            cell = ws.cell(row=r_idx, column=c_idx, value="\n".join(lines))
            cell.alignment = EXPORT_WRAP

    for col_idx in range(1, len(day_list) + 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = EXPORT_HEADER_FILL
        cell.font = EXPORT_HEADER_FONT
        cell.alignment = EXPORT_WRAP
        width = 16 if col_idx > 1 else 12
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions["A"].width = 12

    ws_detail = wb.create_sheet(title="Chi tiet")
    detail_headers = [
        "Ngày",
        "Tiết",
        "Mã môn",
        "Tên môn",
        "Lớp",
        "Giáo viên",
        "Phòng",
        "Loại",
    ]
    for col_idx, h in enumerate(detail_headers, start=1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.fill = EXPORT_HEADER_FILL
        cell.font = EXPORT_HEADER_FONT
    for r_idx, a in enumerate(
        sorted(assigned, key=lambda x: (x["day"], x["period"])), start=2
    ):
        ws_detail.cell(row=r_idx, column=1, value=DAY_NAME_VI.get(a["day"], a["day"]))
        ws_detail.cell(row=r_idx, column=2, value=a["period"] + 1)
        ws_detail.cell(row=r_idx, column=3, value=a["module_code"])
        ws_detail.cell(row=r_idx, column=4, value=a["module_name"])
        ws_detail.cell(row=r_idx, column=5, value=a["group"])
        ws_detail.cell(row=r_idx, column=6, value=a["teachers"])
        ws_detail.cell(row=r_idx, column=7, value=a["room"])
        ws_detail.cell(row=r_idx, column=8, value=a["session_type"])
    for col_idx in range(1, len(detail_headers) + 1):
        ws_detail.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@api.get("/schedule/{schedule_id}/export", auth=tenant_auth)
def export_schedule(request, schedule_id, format: str = "xlsx"):
    schedule = _get_tenant_schedule(request, schedule_id)
    if format != "xlsx":
        raise HttpError(400, "only format=xlsx is supported")
    tenant = request.auth["tenant"]
    payload = _build_export_xlsx(tenant, schedule)
    response = HttpResponse(payload, content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = (
        'attachment; filename="schedule_%s.xlsx"' % schedule_id
    )
    response["Content-Length"] = str(len(payload))
    return response


@api.get("/stats", auth=tenant_auth)
def tenant_stats(request):
    tenant = request.auth["tenant"]
    teachers = tenant.teachers.count()
    student_groups = tenant.student_groups.count()
    resources = tenant.resources.count()
    modules = tenant.modules.count()
    sessions_total = tenant.sessions.count()
    sessions_assigned = tenant.sessions.exclude(
        assigned_timeslot__isnull=True
    ).count()
    loads = {}
    for s in tenant.sessions.exclude(assigned_timeslot__isnull=True):
        ts = s.assigned_timeslot or {}
        d = ts.get("day")
        if d is None:
            continue
        loads[d] = loads.get(d, 0) + 1
    weekly = [
        {"day": DAY_NAME_VI.get(d, str(d)), "load": loads.get(d, 0)}
        for d in sorted(loads)
    ]
    schedules = [
        {
            "id": s.id,
            "name": s.name,
            "status": s.status,
            "objective_value": s.objective_value,
        }
        for s in tenant.schedules.order_by("-id")[:5]
    ]
    completion = 0
    if sessions_total:
        completion = round(sessions_assigned * 100 / sessions_total)
    return {
        "tenant_code": tenant.code,
        "counts": {
            "teachers": teachers,
            "student_groups": student_groups,
            "resources": resources,
            "modules": modules,
            "sessions": sessions_total,
            "sessions_assigned": sessions_assigned,
        },
        "completion": completion,
        "weekly_load": weekly,
        "schedules": schedules,
    }


@api.get("/schedule/{schedule_id}/weights", auth=tenant_auth, response=WeightsOut)
def get_schedule_weights(request, schedule_id):
    schedule = _get_tenant_schedule(request, schedule_id)
    merged = objectives.merge_weights(schedule.weights_json)
    return {"schedule_id": schedule.id, "weights": merged}


@api.put("/schedule/{schedule_id}/weights", auth=tenant_auth, response=WeightsOut)
def update_schedule_weights(request, schedule_id, payload: WeightsIn):
    schedule = _get_tenant_schedule(request, schedule_id)
    incoming = payload.weights or {}
    cleaned = {}
    for k, v in incoming.items():
        if k not in objectives.KNOWN_WEIGHT_KEYS:
            raise HttpError(400, "unknown weight key: %s" % k)
        try:
            iv = int(v)
        except (TypeError, ValueError):
            raise HttpError(400, "invalid weight value for %s: %s" % (k, v))
        if iv < 0:
            raise HttpError(400, "weight must be >= 0: %s" % k)
        cleaned[k] = iv
    stored = dict(schedule.weights_json or {})
    stored.update(cleaned)
    schedule.weights_json = stored
    schedule.save(update_fields=["weights_json"])
    merged = objectives.merge_weights(schedule.weights_json)
    return {"schedule_id": schedule.id, "weights": merged}


class MoveSessionIn(Schema):
    schedule_id: int
    new_day: int
    new_period: int
    new_resource: str | None = None


@api.post("/sessions/{session_id}/move", auth=tenant_auth, url_name="session_move")
def move_session(request, session_id, payload: MoveSessionIn):
    from scheduler.solver.incremental import InfeasibleMove, incremental_resolve

    tenant = request.auth["tenant"]
    try:
        session_id_int = int(session_id)
    except (TypeError, ValueError):
        raise HttpError(404, "session not found")
    if not tenant.sessions.filter(id=session_id_int).exists():
        raise HttpError(404, "session not found")
    schedule = _get_tenant_schedule(request, payload.schedule_id)
    try:
        diff = incremental_resolve(
            schedule,
            session_id_int,
            payload.new_day,
            payload.new_period,
            payload.new_resource,
        )
    except InfeasibleMove as exc:
        raise HttpError(409, str(exc))
    return diff.to_dict()


@api.post(
    "/schedule/{schedule_id}/solve",
    auth=tenant_auth,
    url_name="schedule_solve",
)
def solve_schedule(request, schedule_id):
    from scheduler.tasks import solve_schedule as solve_schedule_task

    schedule = _get_tenant_schedule(request, schedule_id)
    job = SolveJob.objects.create(
        tenant_id=request.auth["tenant_id"],
        schedule=schedule,
        status=SolveJob.Status.QUEUED,
        phase=None,
        progress=0,
    )
    solve_schedule_task.delay(
        str(schedule.id), {"solve_job_id": str(job.id)}
    )
    return JsonResponse(
        {
            "solve_job_id": str(job.id),
            "schedule_id": schedule.id,
            "status": job.status,
        },
        status=202,
    )


class SolveStatusResponse(Schema):
    job_id: str
    schedule_id: int
    status: str
    phase: str | None = None
    progress: int
    objective_value: float | None = None
    error: str
    metrics: dict


@api.get(
    "/solve/{job_id}/status",
    auth=tenant_auth,
    response=SolveStatusResponse,
    url_name="solve_job_status",
)
def solve_job_status(request, job_id):
    from django.core.exceptions import ValidationError

    try:
        job = SolveJob.objects.select_related("schedule").get(id=job_id)
    except (SolveJob.DoesNotExist, ValidationError, ValueError, TypeError):
        raise HttpError(404, "solve job not found")
    if job.tenant_id != request.auth["tenant_id"]:
        raise HttpError(404, "solve job not found")
    return {
        "job_id": str(job.id),
        "schedule_id": job.schedule_id,
        "status": job.status,
        "phase": job.phase,
        "progress": job.progress,
        "objective_value": job.objective_value,
        "error": job.error,
        "metrics": job.metrics_json or {},
    }


XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
SAMPLE_LIMIT = 3


def _summarize(parsed, issues):
    row_counts = {name: len(parsed[name]) for name in excel_parser.SHEET_NAMES}
    sample = {
        name: parsed[name][:SAMPLE_LIMIT] for name in excel_parser.SHEET_NAMES
    }
    errors = [i for i in issues if i["severity"] == validator.SEVERITY_ERROR]
    warnings = [i for i in issues if i["severity"] == validator.SEVERITY_WARNING]
    return {
        "row_counts": row_counts,
        "issues": issues,
        "errors_count": len(errors),
        "warnings_count": len(warnings),
        "sample": sample,
    }


def _read_upload(file: UploadedFile):
    data = file.read()
    if hasattr(file, "seek"):
        try:
            file.seek(0)
        except Exception:
            pass
    if not data:
        raise HttpError(400, "Empty file upload")
    try:
        return excel_parser.parse_bytes(data)
    except Exception as exc:
        raise HttpError(400, "Could not read workbook: %s" % exc)


@api.get("/import/template", auth=tenant_auth, url_name="import_template")
def import_template(request):
    payload = templates.generate_template()
    response = HttpResponse(payload, content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = 'attachment; filename="import_template.xlsx"'
    response["Content-Length"] = str(len(payload))
    return response


@api.post("/import/upload", auth=tenant_auth, url_name="import_upload")
def import_upload(request, file: UploadedFile = File(...)):
    parsed = _read_upload(file)
    issues = validator.validate(parsed)
    return JsonResponse(_summarize(parsed, issues))


@api.post("/import/commit", auth=tenant_auth, url_name="import_commit")
def import_commit(request, file: UploadedFile = File(...)):
    parsed = _read_upload(file)
    issues = validator.validate(parsed)
    if validator.has_errors(issues):
        return JsonResponse(
            {
                "detail": "validation errors",
                "issues": issues,
                "errors_count": sum(
                    1 for i in issues if i["severity"] == validator.SEVERITY_ERROR
                ),
                "created": {},
            },
            status=400,
        )

    tenant = request.auth["tenant"]
    created = _persist(parsed, tenant)
    return JsonResponse({"created": created, "issues": issues}, status=201)


def _persist(parsed, tenant):
    from scheduler.excel_parser import to_int, to_list, to_str

    teachers = {}
    student_groups = {}
    resources = {}
    modules = {}
    created = {
        "teachers": 0,
        "student_groups": 0,
        "resources": 0,
        "modules": 0,
        "teacher_modules": 0,
        "sessions": 0,
    }

    with transaction.atomic():
        for r in parsed["Teachers"]:
            code = to_str(r.get("code"))
            blocks = [b.lower() for b in to_list(r.get("blocks"))]
            quota = None
            if r.get("quota_standard_hours") not in (None, ""):
                quota = to_int(r.get("quota_standard_hours"))
            teacher = Teacher.objects.create(
                tenant=tenant,
                code=code,
                name=to_str(r.get("name")),
                blocks=blocks,
                quota_standard_hours=quota,
            )
            teachers[code] = teacher
            created["teachers"] += 1

        for r in parsed["StudentGroups"]:
            code = to_str(r.get("code"))
            sg = StudentGroup.objects.create(
                tenant=tenant,
                code=code,
                name=to_str(r.get("name")),
                enrollment_type=to_str(r.get("enrollment_type")).lower(),
                size=to_int(r.get("size")) or 0,
            )
            student_groups[code] = sg
            created["student_groups"] += 1

        for r in parsed["Resources"]:
            code = to_str(r.get("code"))
            quantity = to_int(r.get("quantity")) or 0
            available = to_int(r.get("available_quantity"))
            if available is None:
                available = quantity
            resource = Resource.objects.create(
                tenant=tenant,
                code=code,
                name=to_str(r.get("name")),
                type=to_str(r.get("type")).lower(),
                capacity=to_int(r.get("capacity")) or 0,
                quantity=quantity,
                available_quantity=available,
            )
            resources[code] = resource
            created["resources"] += 1

        for r in parsed["Modules"]:
            code = to_str(r.get("code"))
            sg_code = to_str(r.get("student_group"))
            module = Module.objects.create(
                tenant=tenant,
                code=code,
                name=to_str(r.get("name")),
                theory_hours=to_int(r.get("theory_hours")) or 0,
                practice_hours=to_int(r.get("practice_hours")) or 0,
                student_group=student_groups.get(sg_code) if sg_code else None,
            )
            modules[code] = module
            created["modules"] += 1

        for r in parsed["TeacherModule"]:
            teacher = teachers.get(to_str(r.get("teacher_code")))
            module = modules.get(to_str(r.get("module_code")))
            if teacher and module:
                TeacherModule.objects.create(
                    tenant=tenant, teacher=teacher, module=module
                )
                created["teacher_modules"] += 1

        for r in parsed["FixedSessions"]:
            module = modules.get(to_str(r.get("module_code")))
            sg = student_groups.get(to_str(r.get("student_group_code")))
            resource_code = to_str(r.get("resource_code"))
            resource = resources.get(resource_code) if resource_code else None
            if not (module and sg):
                continue
            session = Session.objects.create(
                tenant=tenant,
                module=module,
                student_group=sg,
                session_type=to_str(r.get("session_type")).lower(),
                duration_slots=to_int(r.get("duration_slots")) or 1,
                tier=to_str(r.get("tier")).lower(),
                assigned_resource=resource,
            )
            teacher_codes = to_list(r.get("teacher_codes"))
            if teacher_codes:
                session.assigned_teachers.set(
                    [teachers[c].id for c in teacher_codes if c in teachers]
                )
            created["sessions"] += 1

    return created
