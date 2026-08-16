import datetime as dt

import jwt
from django.test import TestCase, override_settings
from ninja.testing import TestClient

from scheduler.models import Teacher, Tenant


SIGNING_KEY = "dev-insecure-signing-key-change-me"


def mint_token(tenant_id, deployment_id="dep-1", issuer="", audience=""):
    payload = {
        "user_id": "u-1",
        "email": "ops@example.com",
        "tenant_id": tenant_id,
        "deployment_id": deployment_id,
        "exp": dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=1),
        "iat": dt.datetime.now(dt.timezone.utc),
        "sub": "u-1",
    }
    if issuer:
        payload["iss"] = issuer
    if audience:
        payload["aud"] = audience
    kwargs = {"algorithm": "HS256"}
    return jwt.encode(payload, SIGNING_KEY, **kwargs)


class HealthEndpointTests(TestCase):
    def test_health_no_auth(self):
        client = TestClient(__import__("config.api", fromlist=["api"]).api)
        resp = client.get("/health")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), {"status": "ok"})


class JWTAuthTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.tenant_a = Tenant.objects.create(name="Tenant A", code="TA")
        cls.tenant_b = Tenant.objects.create(name="Tenant B", code="TB")
        cls.teacher_a = Teacher.objects.create(
            tenant=cls.tenant_a, code="T-A1", name="Alice"
        )
        cls.teacher_b = Teacher.objects.create(
            tenant=cls.tenant_b, code="T-B1", name="Bob"
        )

    def setUp(self):
        from config.api import api

        self.client = TestClient(api)

    def _bearer(self, tenant_id):
        return {"Authorization": f"Bearer {mint_token(tenant_id)}"}

    def test_me_with_valid_token_tenant_a_returns_200(self):
        resp = self.client.get("/me", headers=self._bearer(self.tenant_a.id))
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        self.assertEqual(body["tenant_id"], self.tenant_a.id)
        self.assertEqual(body["tenant_code"], "TA")
        self.assertEqual(body["tenant_name"], "Tenant A")
        self.assertEqual(body["deployment_id"], "dep-1")

    def test_me_with_unknown_tenant_returns_403(self):
        unknown_id = self.tenant_b.id + 9_999
        resp = self.client.get("/me", headers=self._bearer(unknown_id))
        self.assertEqual(resp.status_code, 403, resp.content)

    def test_me_without_token_returns_401(self):
        resp = self.client.get("/me")
        self.assertEqual(resp.status_code, 401, resp.content)

    def test_me_with_bad_signature_returns_401(self):
        payload = {
            "tenant_id": self.tenant_a.id,
            "deployment_id": "dep-1",
            "exp": dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=1),
        }
        bad = jwt.encode(payload, "wrong-key", algorithm="HS256")
        resp = self.client.get("/me", headers={"Authorization": f"Bearer {bad}"})
        self.assertEqual(resp.status_code, 401, resp.content)

    def test_teachers_scoped_to_tenant_a_excludes_tenant_b(self):
        resp = self.client.get("/teachers", headers=self._bearer(self.tenant_a.id))
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        codes = [row["code"] for row in body]
        self.assertIn("T-A1", codes)
        self.assertNotIn("T-B1", codes)

    def test_teachers_scoped_to_tenant_b_excludes_tenant_a(self):
        resp = self.client.get("/teachers", headers=self._bearer(self.tenant_b.id))
        self.assertEqual(resp.status_code, 200, resp.content)
        codes = [row["code"] for row in resp.json()]
        self.assertIn("T-B1", codes)
        self.assertNotIn("T-A1", codes)

    def test_token_without_tenant_claim_returns_401(self):
        payload = {
            "deployment_id": "dep-1",
            "exp": dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=1),
        }
        tok = jwt.encode(payload, SIGNING_KEY, algorithm="HS256")
        resp = self.client.get("/me", headers={"Authorization": f"Bearer {tok}"})
        self.assertEqual(resp.status_code, 401, resp.content)


@override_settings(JWT_ISSUER="auth.evo.local", JWT_AUDIENCE="scheduler.evo.local")
class JWTIssuerAudienceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.tenant_a = Tenant.objects.create(name="Tenant A", code="TA")

    def setUp(self):
        from config.api import api

        self.client = TestClient(api)

    def test_me_rejects_token_with_wrong_issuer(self):
        payload = {
            "tenant_id": self.tenant_a.id,
            "deployment_id": "dep-1",
            "iss": "someone-else",
            "aud": "scheduler.evo.local",
            "exp": dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=1),
        }
        tok = jwt.encode(payload, SIGNING_KEY, algorithm="HS256")
        resp = self.client.get("/me", headers={"Authorization": f"Bearer {tok}"})
        self.assertEqual(resp.status_code, 401, resp.content)

    def test_me_accepts_token_with_correct_iss_aud(self):
        payload = {
            "tenant_id": self.tenant_a.id,
            "deployment_id": "dep-1",
            "iss": "auth.evo.local",
            "aud": "scheduler.evo.local",
            "exp": dt.datetime.now(dt.timezone.utc) + dt.timedelta(hours=1),
        }
        tok = jwt.encode(payload, SIGNING_KEY, algorithm="HS256")
        resp = self.client.get("/me", headers={"Authorization": f"Bearer {tok}"})
        self.assertEqual(resp.status_code, 200, resp.content)


def _session(
    sid,
    group="LH1",
    duration=1,
    session_type="theory",
    tier="culture",
    size=20,
    module="MH1",
    teachers=(),
    locked=False,
    timeslot=None,
    resource=None,
):
    return {
        "id": sid,
        "code": "%s/%s/%s" % (module, session_type, sid),
        "session_type": session_type,
        "tier": tier,
        "duration_slots": duration,
        "group_size": size,
        "group_code": group,
        "module_code": module,
        "teacher_codes": list(teachers),
        "is_locked": locked,
        "assigned_timeslot": timeslot,
        "assigned_resource_code": resource,
    }


def _resource(code, rtype="theory_room", capacity=40, quantity=1):
    return {
        "id": abs(hash(code)) % 10000,
        "code": code,
        "type": rtype,
        "capacity": capacity,
        "available_quantity": quantity,
    }


def _data(sessions, resources=None, horizon_cfg=None, teacher_module_map=None):
    from scheduler import horizon as horizon_config

    return {
        "horizon": horizon_config.build(horizon_cfg),
        "sessions": sessions,
        "resources": resources if resources is not None else [_resource("P1")],
        "teachers": [],
        "teacher_module_map": teacher_module_map or {},
        "rules": [],
    }


class HorizonConfigTests(TestCase):
    def test_default_is_six_days_of_five_periods(self):
        from scheduler import horizon

        self.assertEqual(horizon.total_slots(None), 30)

    def test_two_sessions_per_day_doubles_capacity(self):
        from scheduler import horizon

        cfg = {"days_per_week": 6, "periods_per_day": 10, "morning_count": 5}
        self.assertEqual(horizon.total_slots(cfg), 60)
        built = horizon.build(cfg)
        self.assertEqual(len(built), 60)
        self.assertEqual(sum(1 for ts in built if ts["is_morning"]), 30)

    def test_validate_rejects_morning_larger_than_day(self):
        from scheduler import horizon

        errors = horizon.validate({"periods_per_day": 4, "morning_count": 6})
        self.assertEqual(len(errors), 1)

    def test_normalize_clamps_out_of_range_values(self):
        from scheduler import horizon

        cfg = horizon.normalize({"periods_per_day": 999, "weeks": 0})
        self.assertEqual(cfg["periods_per_day"], horizon.MAX_PERIODS_PER_DAY)
        self.assertEqual(cfg["weeks"], 1)


class FeasibilityTests(TestCase):
    def test_group_over_capacity_is_blocking(self):
        from scheduler.solver import feasibility

        sessions = [_session(i, duration=4) for i in range(1, 10)]
        issues = feasibility.check(_data(sessions))
        codes = [i["code"] for i in feasibility.blocking(issues)]
        self.assertIn("group_overloaded", codes)

    def test_group_fits_after_widening_horizon(self):
        from scheduler.solver import feasibility

        sessions = [_session(i, duration=4) for i in range(1, 10)]
        data = _data(
            sessions,
            resources=[_resource("P%d" % i) for i in range(6)],
            horizon_cfg={"periods_per_day": 10, "morning_count": 5},
        )
        issues = feasibility.check(data)
        self.assertEqual(feasibility.blocking(issues), [])

    def test_session_longer_than_a_day_is_blocking(self):
        from scheduler.solver import feasibility

        issues = feasibility.check(_data([_session(1, duration=9)]))
        codes = [i["code"] for i in feasibility.blocking(issues)]
        self.assertIn("session_too_long", codes)

    def test_teacher_forced_by_single_eligible_teacher_counts_as_load(self):
        from scheduler.solver import feasibility

        sessions = [
            _session(i, group="LH%d" % i, duration=4, module="MH1")
            for i in range(1, 10)
        ]
        data = _data(
            sessions,
            resources=[_resource("P%d" % i) for i in range(10)],
            teacher_module_map={"MH1": ["GV1"]},
        )
        issues = feasibility.check(data)
        codes = [i["code"] for i in feasibility.blocking(issues)]
        self.assertIn("teacher_overloaded", codes)

    def test_theory_pool_shortage_is_reported(self):
        from scheduler.solver import feasibility

        sessions = [
            _session(i, group="LH%d" % i, duration=5) for i in range(1, 8)
        ]
        data = _data(sessions, resources=[_resource("P1")])
        issues = feasibility.check(data)
        codes = [i["code"] for i in feasibility.blocking(issues)]
        self.assertIn("resource_pool_overloaded", codes)

    def test_module_without_teacher_is_only_a_warning(self):
        from scheduler.solver import feasibility

        issues = feasibility.check(_data([_session(1)]))
        self.assertEqual(feasibility.blocking(issues), [])
        codes = [i["code"] for i in issues]
        self.assertIn("module_without_teacher", codes)

    def test_practice_session_needs_workshop_or_tool_set(self):
        from scheduler.solver import feasibility

        data = _data(
            [_session(1, session_type="practice")],
            resources=[_resource("P1", rtype="theory_room")],
        )
        codes = [i["code"] for i in feasibility.blocking(feasibility.check(data))]
        self.assertIn("no_matching_resource", codes)

    def test_tool_set_without_capacity_fits_any_group(self):
        from scheduler.solver import feasibility

        session = _session(1, session_type="practice", size=48)
        candidates = feasibility.candidate_resources(
            [
                _resource("X1", rtype="workshop", capacity=20),
                _resource("T1", rtype="tool_set", capacity=0, quantity=5),
            ],
            session,
        )
        self.assertEqual(candidates, ["T1"])

    def test_locked_slot_outside_horizon_is_blocking(self):
        from scheduler.solver import feasibility

        session = _session(
            1, locked=True, timeslot={"index": 999}, resource="P1"
        )
        codes = [
            i["code"] for i in feasibility.blocking(feasibility.check(_data([session])))
        ]
        self.assertIn("locked_slot_invalid", codes)


class SolverPreflightTests(TestCase):
    def test_impossible_data_fails_without_calling_solver(self):
        from scheduler.solver import engine

        sessions = [_session(i, duration=4) for i in range(1, 10)]
        result = engine.build_and_solve(_data(sessions), max_time_seconds=1.0)
        self.assertEqual(result.status, engine.STATUS_DATA_INFEASIBLE)
        self.assertFalse(result.is_feasible)
        self.assertTrue(any("Lớp LH1" in v for v in result.violations))
        self.assertEqual(result.wall_time, 0.0)

    def test_feasible_data_still_solves(self):
        from scheduler.solver import engine

        sessions = [
            _session(1, duration=2),
            _session(2, duration=2, group="LH2"),
            _session(3, session_type="practice", group="LH2"),
        ]
        data = _data(
            sessions,
            resources=[
                _resource("P1"),
                _resource("P2"),
                _resource("T1", rtype="tool_set", capacity=0, quantity=2),
            ],
        )
        result = engine.build_and_solve(data, max_time_seconds=10.0)
        self.assertTrue(result.is_feasible, result.violations)
        self.assertEqual(len(result.assignments), 3)

    def test_practice_session_never_lands_in_a_theory_room(self):
        from scheduler.solver import engine

        sessions = [_session(1, session_type="practice", size=10)]
        data = _data(
            sessions,
            resources=[
                _resource("P1", rtype="theory_room"),
                _resource("X1", rtype="workshop", capacity=20),
            ],
        )
        result = engine.build_and_solve(data, max_time_seconds=10.0)
        self.assertTrue(result.is_feasible, result.violations)
        self.assertEqual(result.assignments[0].resource_code, "X1")

    def test_invalid_lock_is_dropped_instead_of_breaking_the_model(self):
        from scheduler.solver import engine

        # Khoá trỏ vào tiết 29 nhưng buổi dài 3 tiết nên sẽ tràn sang ngày sau.
        session = _session(
            1, duration=3, locked=True, timeslot={"index": 29}, resource="P1"
        )
        data = _data([session], resources=[_resource("P1")])
        result = engine.build_and_solve(
            data, max_time_seconds=10.0, skip_preflight=True
        )
        self.assertTrue(result.is_feasible, result.violations)
        self.assertTrue(any("bị khoá" in v for v in result.violations))


class HorizonApiTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.tenant = Tenant.objects.create(name="Tenant H", code="TH")

    def setUp(self):
        from config.api import api

        self.client = TestClient(api)
        self.headers = {"Authorization": f"Bearer {mint_token(self.tenant.id)}"}

    def test_get_returns_default_until_configured(self):
        resp = self.client.get("/tenant/horizon", headers=self.headers)
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        self.assertTrue(body["is_default"])
        self.assertEqual(body["horizon"]["total_slots"], 30)

    def test_put_persists_and_widens_the_week(self):
        resp = self.client.put(
            "/tenant/horizon",
            json={"days_per_week": 6, "periods_per_day": 10, "morning_count": 5},
            headers=self.headers,
        )
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["horizon"]["total_slots"], 60)

        self.tenant.refresh_from_db()
        self.assertEqual(
            self.tenant.config_json["horizon"]["periods_per_day"], 10
        )

    def test_put_rejects_invalid_config(self):
        resp = self.client.put(
            "/tenant/horizon",
            json={"periods_per_day": 4, "morning_count": 9},
            headers=self.headers,
        )
        self.assertEqual(resp.status_code, 400, resp.content)
        self.tenant.refresh_from_db()
        self.assertEqual(self.tenant.config_json, {})


class ExcelConfigSheetTests(TestCase):
    def test_config_sheet_is_optional(self):
        from scheduler import excel_parser

        self.assertIsNone(excel_parser.horizon_config({"Config": []}))
        self.assertNotIn("Config", excel_parser.SHEET_NAMES)

    def test_config_sheet_values_become_horizon(self):
        from scheduler import excel_parser, horizon

        parsed = {
            "Config": [
                {"row": 2, "weeks": 1, "days_per_week": 5, "periods_per_day": 10, "morning_count": 5}
            ]
        }
        cfg = excel_parser.horizon_config(parsed)
        self.assertEqual(horizon.total_slots(cfg), 50)

    def test_template_contains_config_sheet(self):
        import io

        from openpyxl import load_workbook

        from scheduler import templates

        wb = load_workbook(io.BytesIO(templates.generate_template()))
        self.assertIn("Cấu hình", wb.sheetnames)


class FeasibilityEndpointTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from scheduler.models import (
            Module,
            Resource,
            Schedule,
            Session,
            StudentGroup,
        )

        cls.tenant = Tenant.objects.create(name="Tenant F", code="TF")
        group = StudentGroup.objects.create(
            tenant=cls.tenant,
            code="LH1",
            name="Lớp 1",
            enrollment_type=StudentGroup.EnrollmentType.COLLEGE,
            size=20,
        )
        module = Module.objects.create(
            tenant=cls.tenant, code="MH1", name="Môn 1", student_group=group
        )
        Resource.objects.create(
            tenant=cls.tenant,
            code="P1",
            name="Phòng 1",
            type=Resource.ResourceType.THEORY_ROOM,
            capacity=40,
            quantity=1,
            available_quantity=1,
        )
        # 9 buổi x 4 tiết = 36 tiết, vượt 30 tiết của một tuần mặc định.
        for _ in range(9):
            Session.objects.create(
                tenant=cls.tenant,
                module=module,
                student_group=group,
                session_type=Session.SessionType.THEORY,
                duration_slots=4,
                tier=Session.Tier.CULTURE,
            )
        cls.schedule = Schedule.objects.create(
            tenant=cls.tenant, name="Học kỳ 1"
        )

    def setUp(self):
        from config.api import api

        self.client = TestClient(api)
        self.headers = {"Authorization": f"Bearer {mint_token(self.tenant.id)}"}

    def test_reports_the_overloaded_group(self):
        resp = self.client.get(
            "/schedule/%d/feasibility" % self.schedule.id, headers=self.headers
        )
        self.assertEqual(resp.status_code, 200, resp.content)
        body = resp.json()
        self.assertFalse(body["feasible"])
        messages = [
            issue["message"]
            for tier in body["tiers"]
            for issue in tier["issues"]
        ]
        self.assertTrue(
            any("Lớp LH1 cần 36 tiết" in m for m in messages), messages
        )

    def test_widening_the_week_clears_the_blocker(self):
        put = self.client.put(
            "/tenant/horizon",
            json={"days_per_week": 6, "periods_per_day": 10, "morning_count": 5},
            headers=self.headers,
        )
        self.assertEqual(put.status_code, 200, put.content)

        resp = self.client.get(
            "/schedule/%d/feasibility" % self.schedule.id, headers=self.headers
        )
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertTrue(resp.json()["feasible"], resp.json())

    def test_other_tenant_cannot_read_the_report(self):
        other = Tenant.objects.create(name="Tenant G", code="TG")
        resp = self.client.get(
            "/schedule/%d/feasibility" % self.schedule.id,
            headers={"Authorization": f"Bearer {mint_token(other.id)}"},
        )
        self.assertEqual(resp.status_code, 404, resp.content)


class TeacherPoolTests(TestCase):
    def test_two_teachers_sharing_more_slots_than_they_can_teach(self):
        from scheduler.solver import feasibility

        # 7 lớp x 5 tiết = 35 tiết, hai giáo viên chỉ gánh nổi 2 x 30 = 60...
        # nhưng đủ, nên tăng lên 13 lớp = 65 tiết để vượt năng lực.
        sessions = [
            _session(i, group="LH%d" % i, duration=5, module="MH_VAN")
            for i in range(1, 14)
        ]
        data = _data(
            sessions,
            resources=[_resource("P%d" % i) for i in range(20)],
            teacher_module_map={"MH_VAN": ["GV1", "GV2"]},
        )
        codes = [i["code"] for i in feasibility.blocking(feasibility.check(data))]
        self.assertIn("teacher_pool_overloaded", codes)

    def test_pool_with_enough_teachers_is_accepted(self):
        from scheduler.solver import feasibility

        sessions = [
            _session(i, group="LH%d" % i, duration=5, module="MH_VAN")
            for i in range(1, 14)
        ]
        data = _data(
            sessions,
            resources=[_resource("P%d" % i) for i in range(20)],
            teacher_module_map={"MH_VAN": ["GV1", "GV2", "GV3"]},
        )
        self.assertEqual(feasibility.blocking(feasibility.check(data)), [])

    def test_fixed_load_counts_against_the_pool(self):
        from scheduler.solver import feasibility

        # GV1 đã bị gán cứng 28 tiết ở môn khác, nên nhóm hai người chỉ còn
        # 32 tiết trống mà phải gánh 36 tiết môn Văn.
        sessions = [
            _session(
                100 + i,
                group="LHX%d" % i,
                duration=4,
                module="MH_KHAC",
                teachers=("GV1",),
            )
            for i in range(7)
        ]
        sessions += [
            _session(i, group="LH%d" % i, duration=4, module="MH_VAN")
            for i in range(1, 10)
        ]
        data = _data(
            sessions,
            resources=[_resource("P%d" % i) for i in range(20)],
            teacher_module_map={"MH_VAN": ["GV1", "GV2"]},
        )
        codes = [i["code"] for i in feasibility.blocking(feasibility.check(data))]
        self.assertIn("teacher_pool_overloaded", codes)

    def test_single_teacher_pool_is_left_to_the_per_teacher_check(self):
        from scheduler.solver import feasibility

        sessions = [
            _session(i, group="LH%d" % i, duration=5, module="MH_VAN")
            for i in range(1, 9)
        ]
        data = _data(
            sessions,
            resources=[_resource("P%d" % i) for i in range(20)],
            teacher_module_map={"MH_VAN": ["GV1"]},
        )
        codes = [i["code"] for i in feasibility.blocking(feasibility.check(data))]
        self.assertIn("teacher_overloaded", codes)
        self.assertNotIn("teacher_pool_overloaded", codes)
