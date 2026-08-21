import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scheduler.excel_parser import ALL_SHEET_NAMES

SHEET_HEADERS = {
    "Teachers": ["Mã GV", "Họ tên", "Khối", "Định mức"],
    "StudentGroups": ["Mã LH", "Tên LH", "Loại hình", "Sĩ số", "Lớp văn hoá", "Nghề", "Độc hại"],
    "Homerooms": ["Mã lớp", "Tên lớp", "Khối", "Sĩ số", "Ca văn hoá", "Phòng"],
    "Resources": ["Mã TB", "Tên TB", "Loại", "Sức chứa", "Số lượng", "Còn lại"],
    "Modules": ["Mã MH", "Tên MH", "Lý thuyết", "Thực hành", "Mã LH"],
    "TeacherModule": ["Mã GV", "Mã MH"],
    "FixedSessions": ["Mã MH", "Mã LH", "Loại", "Số tiết", "Cấp", "Mã TB", "Mã GV"],
    "Config": [
        "Số tuần",
        "Số ngày học",
        "Số tiết mỗi ngày",
        "Số tiết buổi sáng",
    ],
}

SHEET_LABELS = {
    "Teachers": "Giáo viên",
    "StudentGroups": "Nhóm nghề",
    "Homerooms": "Lớp văn hoá",
    "Resources": "Thiết bị",
    "Modules": "Môn học",
    "TeacherModule": "GV - Môn học",
    "FixedSessions": "Tiết cố định",
    "Config": "Cấu hình",
}

EXAMPLE_ROWS = {
    "Teachers": [
        ["GV001", "Nguyễn Văn An", "Văn hóa", 14],
        ["GV002", "Trần Thị Bình", "Nghề", 16],
    ],
    "StudentGroups": [
        ["G1", "CNKT Điều khiển tự động", "Song bằng", 16, "11A3", "CNKT Điều khiển tự động", ""],
        ["G4", "Thiết kế nội thất", "Song bằng", 17, "11A3", "Thiết kế nội thất", ""],
        ["M3", "KT lắp đặt điện", "Song bằng", 44, "12A1 + 12A4", "KT lắp đặt điện", ""],
        ["LH002", "Lớp 11B2", "Cao đẳng", 25, "", "", ""],
    ],
    "Homerooms": [
        ["11A3", "Lớp 11A3", 11, 63, "Chiều", ""],
        ["12A1", "Lớp 12A1", 12, 30, "Cả ngày", ""],
        ["12A4", "Lớp 12A4", 12, 14, "Cả ngày", ""],
    ],
    "Resources": [
        ["P101", "Phòng Lý thuyết 101", "Phòng lý thuyết", 40, 1, 1],
        ["TS01", "Bộ dụng cụ thực hành", "Bộ dụng cụ", 0, 10, 10],
    ],
    "Modules": [
        ["MH001", "Toán học", 60, 0, "LH001"],
        ["MH002", "Cơ khí", 0, 90, "LH002"],
    ],
    "TeacherModule": [
        ["GV001", "MH001"],
        ["GV002", "MH002"],
    ],
    "FixedSessions": [
        ["MH001", "LH001", "Lý thuyết", 3, "Văn hóa", "P101", "GV001"],
        ["MH002", "LH002", "Thực hành", 4, "Nghề", "TS01", "GV002"],
    ],
    "Config": [[1, 6, 5, 2]],
}

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _write_sheet(wb, sheet_name, headers, examples):
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_data in examples:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[column_letter].width = min(max_len + 4, 40)


def generate_template():
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in ALL_SHEET_NAMES:
        _write_sheet(
            wb,
            SHEET_LABELS[sheet_name],
            SHEET_HEADERS[sheet_name],
            EXAMPLE_ROWS[sheet_name],
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
