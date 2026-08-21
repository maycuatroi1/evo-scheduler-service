import io
import unicodedata

from openpyxl import load_workbook

SHEET_NAMES = [
    "Teachers",
    "StudentGroups",
    "Resources",
    "Modules",
    "TeacherModule",
    "FixedSessions",
]

# Sheet không bắt buộc: thiếu thì giữ nguyên cấu hình đang có, không cảnh báo.
OPTIONAL_SHEET_NAMES = ["Config", "Homerooms"]

ALL_SHEET_NAMES = SHEET_NAMES + OPTIONAL_SHEET_NAMES

SHEET_ALIASES = {
    "Teachers": ["Teachers", "GiaoVien", "Giáo viên", "Giao vien", "GiangVien", "Teacher"],
    "StudentGroups": ["StudentGroups", "LopHoc", "Lớp học", "Lop hoc", "NhomHocSinh", "StudentGroup"],
    "Resources": ["Resources", "ThietBi", "Thiết bị", "Thiet bi", "PhongHoc", "Resource"],
    "Modules": ["Modules", "MonHoc", "Môn học", "Mon hoc", "Module"],
    "TeacherModule": ["TeacherModule", "GVMH", "GV-MH", "GV - Môn học", "GV - Mon hoc", "GV - Môn"],
    "FixedSessions": ["FixedSessions", "TietCoDinh", "Tiết cố định", "Tiet co dinh", "Sessions", "BuoiHoc", "Session"],
    "Config": ["Config", "CauHinh", "Cấu hình", "Cau hinh", "ThoiKhoaBieu", "Thời khoá biểu", "Khung giờ"],
    "Homerooms": [
        "Homerooms",
        "LopVanHoa",
        "Lớp văn hoá",
        "Lop van hoa",
        "Lớp văn hóa",
        "LopVH",
        "Homeroom",
    ],
}

FIELD_ALIASES = {
    "Homerooms": {
        "code": ["Ma lop", "Mã lớp", "Ma LVH", "Mã LVH", "code", "malop"],
        "name": ["Ten lop", "Tên lớp", "name", "tenlop", "ten"],
        "grade": ["Khoi", "Khối", "grade", "khoi", "khoi lop"],
        "size": ["Si so", "Sĩ số", "size", "siso"],
        "culture_shift": [
            "Ca van hoa",
            "Ca văn hoá",
            "Ca văn hóa",
            "culture_shift",
            "cavanhoa",
            "Buoi hoc van hoa",
            "Buổi học văn hoá",
        ],
        "room": ["Phong", "Phòng", "room", "phong", "Ma phong", "Mã phòng"],
    },
    "Teachers": {
        "code": ["Ma GV", "Mã GV", "code", "magv", "codegv", "teacher_code"],
        "name": ["Ho ten", "Họ tên", "name", "hoten", "tengv", "ten"],
        "blocks": ["blocks", "Khoi", "Khối", "khoi", "block"],
        "quota_standard_hours": [
            "quota",
            "Dinh muc",
            "Định mức",
            "dinhmuc",
            "sobuoi",
            "so buoi",
            "quota_standard_hours",
            "standard hours",
        ],
    },
    "StudentGroups": {
        "code": [
            "Ma LH",
            "Mã LH",
            "Ma Lop",
            "Mã Lớp",
            "code",
            "malop",
            "malh",
        ],
        "name": [
            "Ten LH",
            "Tên LH",
            "Ten Lop",
            "Tên Lớp",
            "name",
            "tenlop",
            "tenlh",
            "ten",
        ],
        "enrollment_type": [
            "Loai Hinh",
            "Loại hình",
            "enrollment_type",
            "loaihinh",
            "loai hinh",
        ],
        "size": ["Si so", "Sĩ số", "size", "siso", "soluonghs"],
        # Nhóm nghề gộp nhiều lớp thì liệt kê cách nhau bởi dấu phẩy hoặc
        # dấu cộng: "12A1 + 12A4". Để trống nếu nhóm không thuộc hệ 9+.
        "homeroom_codes": [
            "Lop van hoa",
            "Lớp văn hoá",
            "Lớp văn hóa",
            "homeroom_codes",
            "lopvanhoa",
            "Ma lop VH",
            "Mã lớp VH",
        ],
        "occupation": ["Nghe", "Nghề", "occupation", "nghe", "Ten nghe", "Tên nghề"],
        "hazardous": [
            "Doc hai",
            "Độc hại",
            "hazardous",
            "dochai",
            "Nang nhoc doc hai",
            "Nặng nhọc độc hại",
        ],
    },
    "Resources": {
        "code": [
            "Ma TB",
            "Mã TB",
            "Ma Phong",
            "Mã Phòng",
            "code",
            "matb",
            "maphong",
        ],
        "name": [
            "Ten TB",
            "Tên TB",
            "Ten Phong",
            "Tên Phòng",
            "name",
            "tentb",
            "tenphong",
            "ten",
        ],
        "type": ["Loai", "Loại", "type", "loai", "loaitb"],
        "capacity": ["Suc chua", "Sức chứa", "capacity", "succhua"],
        "quantity": ["So luong", "Số lượng", "quantity", "soluong", "sl"],
        "available_quantity": [
            "Con lai",
            "Còn lại",
            "available",
            "available_quantity",
            "conlai",
        ],
    },
    "Modules": {
        "code": ["Ma MH", "Mã MH", "code", "mamh", "module_code"],
        "name": ["Ten MH", "Tên MH", "name", "tenmh", "ten"],
        "theory_hours": [
            "LT",
            "Ly thuyet",
            "Lý thuyết",
            "theory_hours",
            "lythuyet",
            "so tiet lt",
        ],
        "practice_hours": [
            "TH",
            "Thuc hanh",
            "Thực hành",
            "practice_hours",
            "thuchanh",
            "so tiet th",
        ],
        "student_group": [
            "Ma LH",
            "Mã LH",
            "student_group",
            "malh",
            "lop",
            "student_group_code",
        ],
    },
    "TeacherModule": {
        "teacher_code": [
            "Ma GV",
            "Mã GV",
            "teacher",
            "teacher_code",
            "magv",
        ],
        "module_code": [
            "Ma MH",
            "Mã MH",
            "module",
            "module_code",
            "mamh",
        ],
    },
    "FixedSessions": {
        "module_code": [
            "Ma MH",
            "Mã MH",
            "module",
            "module_code",
            "mamh",
        ],
        "student_group_code": [
            "Ma LH",
            "Mã LH",
            "student_group",
            "student_group_code",
            "malh",
        ],
        "session_type": [
            "Loai",
            "Loại",
            "session_type",
            "loai",
            "loaibt",
            "type",
        ],
        "duration_slots": [
            "So tiet",
            "Số tiết",
            "duration_slots",
            "slot",
            "sotiet",
            "slots",
        ],
        "tier": ["Cap", "Cấp", "tier", "cap", "loaicap"],
        "resource_code": [
            "Ma TB",
            "Mã TB",
            "resource",
            "resource_code",
            "matb",
            "phong",
        ],
        "teacher_codes": [
            "Ma GV",
            "Mã GV",
            "teacher",
            "teachers",
            "teacher_codes",
            "magv",
        ],
    },
    "Config": {
        "weeks": ["So tuan", "Số tuần", "weeks", "sotuan"],
        "days_per_week": [
            "So ngay hoc",
            "Số ngày học",
            "So ngay",
            "Số ngày",
            "days_per_week",
            "songay",
        ],
        "periods_per_day": [
            "So tiet moi ngay",
            "Số tiết mỗi ngày",
            "So tiet/ngay",
            "Số tiết/ngày",
            "periods_per_day",
            "sotietngay",
        ],
        "morning_count": [
            "So tiet buoi sang",
            "Số tiết buổi sáng",
            "morning_count",
            "sotietsang",
        ],
    },
}


def normalize(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    decomposed = unicodedata.normalize("NFKD", text)
    ascii_only = "".join(c for c in decomposed if not unicodedata.combining(c))
    return "".join(ascii_only.split())


def to_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_int(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            return None
        return int(value)
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def to_list(value):
    if value is None or value == "":
        return []
    if isinstance(value, (list, tuple)):
        return [to_str(v) for v in value if to_str(v)]
    if isinstance(value, str):
        # Thời khoá biểu của trường ghi nhóm gộp là "12A1 + 12A4", nên
        # tách cả dấu cộng chứ không chỉ dấu phẩy.
        parts = value.replace(";", ",").replace("+", ",").split(",")
        return [p.strip() for p in parts if p.strip()]
    return [to_str(value)]


ENUM_LABELS = {
    "morning": ["Sáng", "Buổi sáng", "Ca sáng"],
    "afternoon": ["Chiều", "Buổi chiều", "Ca chiều"],
    "full_day": ["Cả ngày", "Cả hai buổi", "Sáng và chiều"],
    "culture": ["Văn hóa", "Khối văn hóa"],
    "vocational": ["Nghề", "Khối nghề"],
    "both": ["Cả hai"],
    "dual_degree": ["Song bằng"],
    "college": ["Cao đẳng"],
    "theory_room": ["Phòng lý thuyết", "Phòng học lý thuyết"],
    "tool_set": ["Bộ dụng cụ", "Dụng cụ thực hành"],
    "workshop": ["Xưởng"],
    "theory": ["Lý thuyết"],
    "practice": ["Thực hành"],
}

ENUM_NORMALIZED = {}
for _code, _labels in ENUM_LABELS.items():
    for _lbl in [_code, _code.replace("_", ""), _code.replace("_", " ")] + _labels:
        ENUM_NORMALIZED[normalize(_lbl)] = _code

ENUM_FIELDS = (
    "blocks",
    "enrollment_type",
    "type",
    "session_type",
    "tier",
    "culture_shift",
)


def _canonical_enum(raw):
    text = to_str(raw)
    if not text:
        return raw
    return ENUM_NORMALIZED.get(normalize(text), text)


def _match_sheets(workbook):
    canonical_by_norm = {}
    for canonical, aliases in SHEET_ALIASES.items():
        canonical_by_norm[normalize(canonical)] = canonical
        for alias in aliases:
            canonical_by_norm[normalize(alias)] = canonical
    mapping = {}
    for sheet_name in workbook.sheetnames:
        norm = normalize(sheet_name)
        if norm and norm in canonical_by_norm:
            mapping.setdefault(canonical_by_norm[norm], sheet_name)
    return mapping


def _match_headers(header_row, aliases):
    norm_to_field = {}
    for field, alias_list in aliases.items():
        norm_to_field[normalize(field)] = field
        for alias in alias_list:
            norm_to_field[normalize(alias)] = field
    field_to_col = {}
    for col_idx, cell_value in enumerate(header_row, start=1):
        norm = normalize(cell_value)
        if norm and norm in norm_to_field:
            field = norm_to_field[norm]
            if field not in field_to_col:
                field_to_col[field] = col_idx
    return field_to_col


def _read_rows(ws, field_to_col):
    rows = []
    max_col = ws.max_column or 0
    for row_idx in range(2, (ws.max_row or 1) + 1):
        cells = [
            ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)
        ]
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in cells):
            continue
        record = {"row": row_idx}
        for field, col_idx in field_to_col.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if field == "blocks":
                val = [_canonical_enum(v) for v in to_list(val)]
            elif field in ENUM_FIELDS:
                val = _canonical_enum(val)
            record[field] = val
        rows.append(record)
    return rows


def parse_workbook(file_obj_or_path):
    if hasattr(file_obj_or_path, "read"):
        wb = load_workbook(file_obj_or_path, data_only=True)
    else:
        wb = load_workbook(filename=file_obj_or_path, data_only=True)

    sheet_map = _match_sheets(wb)
    result = {}
    missing = []
    for canonical in ALL_SHEET_NAMES:
        if canonical not in sheet_map:
            result[canonical] = []
            if canonical not in OPTIONAL_SHEET_NAMES:
                missing.append(canonical)
            continue
        ws = wb[sheet_map[canonical]]
        field_to_col = {}
        if ws.max_row and ws.max_row >= 1:
            header_row = [
                ws.cell(row=1, column=c).value
                for c in range(1, (ws.max_column or 0) + 1)
            ]
            field_to_col = _match_headers(header_row, FIELD_ALIASES[canonical])
        result[canonical] = _read_rows(ws, field_to_col)
    result["_missing_sheets"] = missing
    return result


def parse_bytes(data):
    return parse_workbook(io.BytesIO(data))


CONFIG_FIELDS = ("weeks", "days_per_week", "periods_per_day", "morning_count")


def horizon_config(parsed):
    """Cấu hình khung giờ khai trong sheet Config, hoặc None nếu bỏ trống.

    Chỉ đọc dòng đầu tiên: khung giờ là thiết lập của cả trường, không phải
    danh sách.
    """
    rows = parsed.get("Config") or []
    for row in rows:
        cfg = {}
        for field in CONFIG_FIELDS:
            value = to_int(row.get(field))
            if value is not None:
                cfg[field] = value
        if cfg:
            return cfg
    return None
