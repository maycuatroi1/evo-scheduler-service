"""
Generate a large realistic Excel file for evo-scheduler-service demo.
~1500 students, 50 teachers, 35 resources, 40 modules, ~500 sessions.
Two-tier: culture (semester-locked) + vocational (weekly).
"""
import io
import random
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

random.seed(2026)

LAST_NAMES = [
    "Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Phan", "Vũ", "Võ",
    "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương", "Lý",
]
MALE_NAMES = [
    "Văn An", "Văn Bình", "Văn Cường", "Văn Đức", "Văn Đông", "Văn Hà",
    "Văn Hùng", "Văn Khánh", "Văn Long", "Văn Minh", "Văn Nam", "Văn Phong",
    "Văn Quân", "Văn Sơn", "Văn Thành", "Văn Trung", "Văn Tú", "Văn Dương",
    "Hoàng Nam", "Hoàng Long", "Hoàng Phong", "Hoàng Hải", "Đức Anh", "Minh Tuấn",
    "Quang Huy", "Thanh Hải", "Công Minh", "Tiến Dũng", "Bảo Khánh", "Trọng Nhân",
]
FEMALE_NAMES = [
    "Thị Bình", "Thị Châu", "Thị Dung", "Thị Hà", "Thị Hằng", "Thị Hồng",
    "Thị Lan", "Thị Liễu", "Thị Mai", "Thị Ngọc", "Thị Phương", "Thị Quỳnh",
    "Thị Sen", "Thị Thanh", "Thị Thu", "Thị Trâm", "Thị Vân", "Thị Xuân",
    "Thị Yến", "Thị Hoa", "Thị Hương", "Thị Loan", "Thị Mỹ", "Thị Nga",
    "Thị Oanh", "Thị Phúc", "Thị Tâm", "Thị Uyên", "Thị Vy", "Thị Diệu",
]
DEPARTMENTS = ["Điện", "Cơ khí", "Xây dựng", "Điện tử", "Công nghệ ô tô"]
TRADE_MODULES_CULTURE = [
    ("MH_TOAN", "Toán học", 60, 0),
    ("MH_LY", "Vật lý", 45, 0),
    ("MH_HOA", "Hóa học", 45, 0),
    ("MH_VAN", "Ngữ văn", 60, 0),
    ("MH_SU", "Lịch sử", 30, 0),
    ("MH_ANH", "Tiếng Anh", 60, 0),
    ("MH_CT", "Chính trị", 30, 0),
    ("MH_GDCD", "Giáo dục công dân", 30, 0),
]
TRADE_MODULES_VOCATIONAL = [
    ("MH_DIENCOBAN", "Kỹ thuật điện cơ bản", 30, 60),
    ("MH_DIENCONGNGHIEP", "Lắp đặt điện công nghiệp", 15, 90),
    ("MH_VDK", "Lập trình vi điều khiển", 30, 70),
    ("MH_COKHI", "Kỹ thuật cơ khí", 20, 80),
    ("MH_HAN", "Kỹ thuật hàn", 15, 85),
    ("MH_XAYDUNG", "Kỹ thuật xây dựng", 30, 60),
    ("NH_THO", "Thợ xây cơ bản", 20, 80),
    ("NH_QUYHOACH", "Quy hoạch và thiết kế", 30, 30),
    ("MH_OTO", "Kỹ thuật sửa chữa ô tô", 20, 80),
    ("MH_DIENTU", "Kỹ thuật điện tử", 25, 65),
    ("MH_CAD", "CAD/CAM cơ khí", 15, 45),
    ("NH_NGUONDEN", "Nguồn điện và chiếu sáng", 20, 40),
    ("MH_BAO_TRI", "Bảo trì thiết bị điện", 15, 75),
    ("NH_LAP_DAT", "Lắp đặt thiết bị công nghiệp", 10, 90),
    ("MH_ATLD", "An toàn lao động", 15, 15),
    ("MH_DO_LUONG", "Đo lường và điều khiển", 20, 50),
    ("NH_MAY_DIEU_KHIEN", "Máy điều khiển số CNC", 15, 60),
    ("MH_HE_THONG_DIEN", "Hệ thống điện thông minh", 20, 55),
]
WORKSHOPS_BY_DEPT = {
    "Điện": [
        ("XN_DIEN1", "Xưởng điện công nghiệp 1", "Xưởng", 20),
        ("XN_DIEN2", "Xưởng điện công nghiệp 2", "Xưởng", 15),
        ("XN_VDK", "Xưởng vi điều khiển", "Xưởng", 12),
    ],
    "Cơ khí": [
        ("XN_COKHI1", "Xưởng cơ khí 1", "Xưởng", 20),
        ("XN_COKHI2", "Xưởng hàn - gò", "Xưởng", 15),
        ("XN_CNC", "Xưởng CNC", "Xưởng", 8),
    ],
    "Xây dựng": [
        ("XN_XD1", "Xưởng thí nghiệm vật liệu", "Xưởng", 20),
        ("XN_XD2", "Xưởng thực hành thợ xây", "Xưởng", 15),
    ],
    "Điện tử": [
        ("XN_DT1", "Xưởng điện tử 1", "Xưởng", 20),
        ("XN_DT2", "Xưởng điện tử 2", "Xưởng", 12),
    ],
    "Công nghệ ô tô": [
        ("XN_OTO1", "Xưởng sửa chữa ô tô", "Xưởng", 10),
        ("XN_OTO2", "Xưởng động cơ", "Xưởng", 8),
    ],
}
TOOL_SETS = [
    ("TS_DIEN_01", "Bộ dụng cụ điện A", "Bộ dụng cụ", 5),
    ("TS_DIEN_02", "Bộ dụng cụ điện B", "Bộ dụng cụ", 5),
    ("TS_DIEN_03", "Bộ đo điện đa năng", "Bộ dụng cụ", 8),
    ("TS_COKHI_01", "Bộ dụng cụ cơ khí A", "Bộ dụng cụ", 6),
    ("TS_COKHI_02", "Bộ dụng cụ cơ khí B", "Bộ dụng cụ", 6),
    ("TS_HAN_01", "Bộ dụng cụ hàn", "Bộ dụng cụ", 4),
    ("TS_HAN_02", "Máy hàn hồ quang", "Bộ dụng cụ", 3),
    ("TS_XD_01", "Bộ dụng cụ xây dựng A", "Bộ dụng cụ", 8),
    ("TS_XD_02", "Bộ dụng cụ xây dựng B", "Bộ dụng cụ", 8),
    ("TS_OTO_01", "Bộ dụng cụ ô tô", "Bộ dụng cụ", 5),
    ("TS_VDK_01", "Board Arduino + linh kiện", "Bộ dụng cụ", 10),
    ("TS_VDK_02", "Board mô phỏng PLC", "Bộ dụng cụ", 6),
    ("TS_DT_01", "Bộ dụng cụ điện tử A", "Bộ dụng cụ", 8),
    ("TS_DT_02", "Máy hiện sóng", "Bộ dụng cụ", 4),
]

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def make_name(gender=None):
    ln = random.choice(LAST_NAMES)
    if gender is None:
        gender = random.choice(["M", "F"])
    fn = random.choice(MALE_NAMES if gender == "M" else FEMALE_NAMES)
    return f"{ln} {fn}"


def build_data():
    teachers = []
    student_groups = []
    resources = []
    modules = []
    teacher_modules = []
    sessions = []
    teacher_map = {}

    gv_counter = 0
    for dept in DEPARTMENTS:
        for _ in range(8):
            gv_counter += 1
            code = f"GV{gv_counter:03d}"
            name = make_name()
            dept_lower = dept.lower()
            blocks = ["Nghề"]
            t = (code, name, "Nghề", None, dept)
            teachers.append(t)
            teacher_map[code] = dept

    culture_subject_teachers = {}
    for i, (mcode, mname, _, _) in enumerate(TRADE_MODULES_CULTURE):
        for j in range(2):
            gv_counter += 1
            code = f"GV{gv_counter:03d}"
            name = make_name()
            teachers.append((code, name, "Văn hóa", None, "__culture__"))
            teacher_map[code] = "__culture__"
            culture_subject_teachers.setdefault(mcode, []).append(code)

    for i in range(5):
        gv_counter += 1
        code = f"GV{gv_counter:03d}"
        name = make_name()
        teachers.append((code, name, "Cả hai", None, "Cả hai"))
        teacher_map[code] = "Cả hai"

    theory_rooms = []
    for i in range(1, 16):
        code = f"P{i:03d}"
        cap = random.choice([30, 35, 40, 45, 50])
        theory_rooms.append(code)
        resources.append((code, f"Phòng lý thuyết {i:03d}", "Phòng lý thuyết", cap, 1, 1))

    for dept, ws_list in WORKSHOPS_BY_DEPT.items():
        for code, name, rtype, cap in ws_list:
            resources.append((code, name, rtype, cap, 1, 1))

    for code, name, rtype, qty in TOOL_SETS:
        resources.append((code, name, "Bộ dụng cụ", 0, qty, qty))

    resource_codes = {r[0] for r in resources}
    tool_by_keyword = {}
    for code, name, _, _, _, _ in resources:
        for kw in ["điện", "cơ khí", "hàn", "xây", "ô tô", "arduino", "plc", "điện tử", "hiện sóng"]:
            if kw in name.lower():
                tool_by_keyword.setdefault(kw, []).append(code)

    workshop_by_dept = {}
    for code, name, rtype, _, _, _ in resources:
        if rtype == "Xưởng":
            for dept in DEPARTMENTS:
                dept_kw = dept.lower()
                if dept_kw in name.lower() or (
                    dept == "Điện" and "điện" in name.lower() and "tử" not in name.lower()
                ) or (dept == "Điện tử" and "điện tử" in name.lower()):
                    workshop_by_dept.setdefault(dept, []).append(code)

    culture_classes = []
    for i in range(15):
        code = f"VH{i+1:02d}"
        size = random.randint(38, 48)
        name = f"Lớp Văn hóa {i+1:02d}"
        student_groups.append((code, name, "Cao đẳng", size))
        culture_classes.append(code)

    vocational_classes = []
    for dept in DEPARTMENTS:
        n = random.randint(5, 7)
        for i in range(n):
            idx = len(vocational_classes) + 1
            code = f"NG{idx:02d}"
            size = random.choice([4, 5, 6, 8, 10, 12, 15, 18, 20, 22, 25])
            name = f"Lớp Nghề {dept} {i+1}"
            student_groups.append((code, name, "Cao đẳng", size))
            vocational_classes.append((code, dept))

    dual_degree_classes = []
    for i in range(10):
        code = f"SB{i+1:02d}"
        size = random.randint(35, 42)
        name = f"Lớp Song bằng {i+1:02d}"
        student_groups.append((code, name, "Song bằng", size))
        dual_degree_classes.append(code)

    total_students = sum(r[3] for r in student_groups)

    all_modules = []
    for mcode, mname, lt, th in TRADE_MODULES_CULTURE:
        for cls in culture_classes + dual_degree_classes:
            cm = (f"{mcode}_{cls}", f"{mname} - {cls}", lt, th, cls, "Văn hóa")
            modules.append(cm)
            all_modules.append(cm)
    for mcode, mname, lt, th in TRADE_MODULES_VOCATIONAL:
        for cls, dept in vocational_classes:
            dept_kw = dept.lower()
            skip = False
            if dept_kw not in mcode.lower() and dept_kw not in mname.lower():
                generic = ["ATLD", "CAD"]
                if not any(g in mcode for g in generic):
                    skip = True
            if skip:
                continue
            vm = (f"{mcode}_{cls}", f"{mname} - {cls}", lt, th, cls, "Nghề")
            modules.append(vm)
            all_modules.append(vm)

    for m in all_modules:
        mcode_full, mname, _, _, cls, tier = m
        if tier == "Văn hóa":
            base_code = mcode_full.rsplit("_", 1)[0]
            candidates = culture_subject_teachers.get(base_code, [])
            if candidates:
                tc = random.choice(candidates)
                teacher_modules.append((tc, mcode_full))
        else:
            base_code = mcode_full.rsplit("_", 1)[0]
            dept_of_class = None
            for cc, cd in vocational_classes:
                if cc == cls:
                    dept_of_class = cd
                    break
            dept_teachers = [
                t[0] for t in teachers if t[4] == dept_of_class
            ] if dept_of_class else []
            if dept_teachers:
                tc = random.choice(dept_teachers)
                teacher_modules.append((tc, mcode_full))
                if random.random() < 0.25:
                    tc2 = random.choice(dept_teachers)
                    if tc2 != tc:
                        teacher_modules.append((tc2, mcode_full))

    session_id = 0
    for m in all_modules:
        mcode_full, mname, lt_hours, th_hours, cls, tier = m
        if lt_hours > 0:
            n_theory = max(1, lt_hours // 25)
            for _ in range(n_theory):
                session_id += 1
                room = random.choice(theory_rooms)
                base_code = mcode_full.rsplit("_", 1)[0]
                tc = None
                for tcode, tmcode in teacher_modules:
                    if tmcode == mcode_full:
                        tc = tcode
                        break
                sessions.append((
                    mcode_full, cls, "Lý thuyết", random.choice([2, 3]),
                    tier, room, tc or "",
                ))
        if th_hours > 0:
            n_practice = max(1, th_hours // 35)
            dept_of_class = None
            for cc, cd in vocational_classes:
                if cc == cls:
                    dept_of_class = cd
                    break
            ws_candidates = workshop_by_dept.get(dept_of_class, [])
            if not ws_candidates:
                ws_candidates = [r[0] for r in resources if r[2] == "Xưởng"]
            tool_candidates = []
            for kw, codes in tool_by_keyword.items():
                if dept_of_class and kw in dept_of_class.lower():
                    tool_candidates.extend(codes)
            if not tool_candidates:
                tool_candidates = [r[0] for r in resources if "Bộ dụng cụ" in r[2]]
            for _ in range(n_practice):
                session_id += 1
                ws = random.choice(ws_candidates) if ws_candidates else ""
                tc = None
                for tcode, tmcode in teacher_modules:
                    if tmcode == mcode_full:
                        tc = tcode
                        break
                tc2 = None
                matching = [t for t in teacher_modules if t[1] == mcode_full]
                if len(matching) > 1:
                    tc2 = matching[1][0]
                teacher_str = tc or ""
                if tc2:
                    teacher_str = f"{tc},{tc2}"
                sessions.append((
                    mcode_full, cls, "Thực hành", random.choice([3, 4]),
                    tier, ws, teacher_str,
                ))

    return teachers, student_groups, resources, modules, teacher_modules, sessions, total_students


def write_sheet(ws, headers, rows):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def generate(filepath):
    teachers, groups, resources, modules, tm, sessions, total_students = build_data()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Giáo viên")
    write_sheet(ws1, ["Mã GV", "Họ tên", "Khối", "Định mức"],
                [(t[0], t[1], t[2], t[3] if t[3] else random.choice([12, 14, 16, 18, 20]))
                 for t in teachers])

    ws2 = wb.create_sheet("Lớp học")
    write_sheet(ws2, ["Mã LH", "Tên LH", "Loại hình", "Sĩ số"], groups)

    ws3 = wb.create_sheet("Thiết bị")
    write_sheet(ws3, ["Mã TB", "Tên TB", "Loại", "Sức chứa", "Số lượng", "Còn lại"],
                [(r[0], r[1], r[2], r[3], r[4], r[5]) for r in resources])

    ws4 = wb.create_sheet("Môn học")
    write_sheet(ws4, ["Mã MH", "Tên MH", "Lý thuyết", "Thực hành", "Mã LH"],
                [(m[0], m[1], m[2], m[3], m[4]) for m in modules])

    ws5 = wb.create_sheet("GV - Môn học")
    write_sheet(ws5, ["Mã GV", "Mã MH"], tm)

    ws6 = wb.create_sheet("Tiết cố định")
    write_sheet(ws6, ["Mã MH", "Mã LH", "Loại", "Số tiết", "Cấp", "Mã TB", "Mã GV"], sessions)

    wb.save(filepath)
    return {
        "teachers": len(teachers),
        "student_groups": len(groups),
        "total_students": total_students,
        "resources": len(resources),
        "modules": len(modules),
        "teacher_modules": len(tm),
        "sessions": len(sessions),
    }


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "demo_large.xlsx"
    stats = generate(out)
    print(f"Generated: {out}")
    for k, v in stats.items():
        print(f"  {k}: {v}")
